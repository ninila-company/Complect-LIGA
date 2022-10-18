import csv
import datetime

from bs4 import BeautifulSoup
from ckeditor_uploader.widgets import CKEditorUploadingWidget
from django import forms
from django.contrib import admin
from django.http import HttpResponse
from django.urls import reverse
from django.utils.safestring import mark_safe
from openpyxl import load_workbook

from .models import (Customer, Equipment, Manager, Order, PaymentType,
                     Postprint, ProductType)


admin.site.register(Postprint)


class OrderAdminForm(forms.ModelForm):
    description = forms.CharField(label='Описание', widget=CKEditorUploadingWidget)

    class Meta:
        model = Order
        fields = '__all__'


@admin.register(Manager)
class ManagerAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'phoneNumber')


@admin.register(Equipment)
class EquipmentAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(PaymentType)
class PaymentTypeAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(ProductType)
class ProductTypeAdmin(admin.ModelAdmin):
    list_display = ('name',)


@admin.register(Customer)
class CastomerAdmit(admin.ModelAdmin):
    list_display = ('name', 'email', 'phoneNumber')


# экспорт данных в xlsx файл
def admin_order_xlsx(modeladmin, request, queryset):
    """Наряд заказ."""
    opts = modeladmin.model._meta

    # если из докера то добавь /app/orders
    fn = 'static/listovki.xlsx'
    # fn = 'static/smart_order.xlsx'
    wb = load_workbook(fn)

    fields = [field for field in opts.get_fields()]
    list_value = []
    for obj in queryset:
        for field in fields:
            value = getattr(obj, field.name)
            list_value.append(value)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={list_value[3]}.xlsx'

    ws = wb['list1']  # номер листа в документе
    ws['G8'] = list_value[1]  # номер заказа
    ws['C15'] = list_value[2]  # название заказа
    ws['C13'] = str(list_value[4])  # тип продукции

    soup = BeautifulSoup(list_value[5], 'html.parser')
    ws['C48'] = soup.get_text()  # описание(Убираем все теги html)

    ws['B11'] = str(list_value[7])  # заказчик
    ws['B17'] = f'{list_value[8]} шт'  # тираж
    ws['B73'] = f'{list_value[8]} шт'  # тираж
    ws['I22'] = str(list_value[9])  # оборудование
    ws['I8'] = str(list_value[10])  # дата принятия заказа
    ws['D3'] = str(list_value[10])  # дата принятия заказа
    ws['I48'] = str(list_value[11])  # дата сдачи заказа
    ws['G77'] = str(list_value[11])  # дата сдачи заказа
    ws['D73'] = str(list_value[11])  # дата сдачи заказа
    ws['B85'] = str(list_value[12])  # менеджер
    ws['B5'] = list_value[13]  # сумма договора
    ws['D5'] = str(list_value[15])  # вид платежа
    ws['B53'] = list_value[17]  # ссылка

    list_postprint = [str(i) for i in list_value[19].all()]
    number_cell = 0
    if len(list_postprint) < 4:
        list_postprint.extend([''] * (4 - len(list_postprint)))
    for i in list_postprint:
        while True:
            ws[f'C4{number_cell}'] = str(i)  # постпечать
            break
        number_cell += 2

    wb.save(fn)
    wb.save(response)
    wb.close()

    return response


admin_order_xlsx.short_description = 'Наряд заказ'


# экспорт данных в csv файл
def export_to_csv(modeladmin, request, queryset):
    """Генерация csv файла."""
    opts = modeladmin.model._meta
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=list-orders.csv'
    writer = csv.writer(response)
    fields = [field for field in opts.get_fields()]
    # заголовки колонок
    header = []
    for field in fields:
        header.append(field.verbose_name)
    header = header[1:3] + header[4:5] + header[6:16]  # выбор нужных полей
    writer.writerow(header)
    # запись данных в колонки
    for obj in queryset:
        data_row = []
        for field in fields:
            value = getattr(obj, field.name)
            if isinstance(value, datetime.datetime):
                value = value.strftime('%d/%m/%Y')
            data_row.append(value)
        data_row = data_row[1:3] + data_row[4:5] + data_row[6:16]  # выбор нужных полей
        writer.writerow(data_row)
    return response


export_to_csv.short_description = 'Экспорт в файл CSV'


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ('name', 'number_order', 'year', 'customer', 'product_type', 'circulation',
                    'equipment', 'date_of_acceptance_of_the_order', 'date_of_delivery_of_the_order',
                    'manager', 'the_amount_of_the_deal', 'the_date_of_payment', 'payment_type',
                    'hypperlink', 'readiness', 'completeness', 'order_pdf')
    prepopulated_fields = {'slug': ('name',)}
    list_filter = ('manager', 'date_of_acceptance_of_the_order', 'customer', 'equipment',
                   'product_type')
    filter_horizontal = ('postprint',)
    list_editable = ('manager', 'readiness', 'completeness')
    save_on_top = True
    form = OrderAdminForm
    actions = [export_to_csv, admin_order_xlsx]

    def order_pdf(self, obj):
        return mark_safe('<a href="{}">PDF</a>'.format(
            reverse('app_order:admin_order_pdf', args=[obj.id])))

    order_pdf.short_description = 'Печать'
